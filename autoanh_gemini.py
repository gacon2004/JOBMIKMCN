# -*- coding: utf-8 -*-
"""
SHOPEE AUTO PICK - GEMINI AI VERSION
Chọn ảnh tốt nhất từ Shopee cho video Veo3 bằng Gemini Vision API
"""
import os, time, glob, subprocess, threading, json, sys
from pathlib import Path
import hashlib
import platform
import uuid
import requests

# GUI / Image
import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk
import cv2
import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
import pyautogui

# Tắt tính năng fail-safe của PyAutoGUI (không báo lỗi khi chuột ra góc màn hình)
pyautogui.FAILSAFE = False

# Gemini AI
import google.generativeai as genai
from PIL import Image as PILImage

# NSFW Detector
from nsfw_detector import predict

# ========= LICENSE SYSTEM - GOOGLE SHEETS =========
GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/17Q_J2X1q3-y1oJdZq8K2E19ATsv6L8156js-uSyw0EQ/export?format=csv"

def _get_machine_id():
    """Lấy Machine ID duy nhất của máy"""
    info = f"{platform.node()}|{platform.machine()}|{uuid.getnode()}"
    return hashlib.sha256(info.encode()).hexdigest()[:16]

def check_license_from_google_sheet():
    """
    Kiểm tra Machine ID có trong Google Sheet không
    Sheet phải có cột đầu tiên là 'key' chứa danh sách Machine ID hợp lệ
    """
    machine_id = _get_machine_id()
    
    # Tạo root window tạm thời cho messagebox
    root = tk.Tk()
    root.withdraw()  # Ẩn cửa sổ chính
    
    try:
        print(f"🔍 Đang kiểm tra license...")
        print(f"   Machine ID: {machine_id}")
        
        # Đọc Google Sheet (export CSV)
        response = requests.get(GOOGLE_SHEET_URL, timeout=10)
        response.raise_for_status()
        
        # Parse CSV
        lines = response.text.strip().split('\n')
        if len(lines) < 2:
            raise ValueError("Google Sheet trống hoặc không đúng định dạng")
        
        # Bỏ qua header, đọc danh sách key
        valid_keys = []
        for line in lines[1:]:  # Skip header
            key = line.strip().strip('"').strip()
            if key:
                valid_keys.append(key)
        
        print(f"   Tìm thấy {len(valid_keys)} key trong database")
        
        # Kiểm tra machine ID
        if machine_id in valid_keys:
            print(f"✅ License hợp lệ!")
            messagebox.showinfo(
                "License Hợp Lệ",
                f"✅ Tool đã được kích hoạt!\n\n"
                f"Machine ID: {machine_id}\n\n"
                f"Đã xác thực với database."
            )
            root.destroy()
            return True
        else:
            print(f"❌ Machine ID không có trong danh sách!")
            messagebox.showerror(
                "License Không Hợp Lệ",
                f"❌ TOOL CHƯA ĐƯỢC KÍCH HOẠT!\n\n"
                f"Machine ID của máy này:\n{machine_id}\n\n"
                f"Vui lòng:\n"
                f"1. Copy Machine ID này\n"
                f"2. Gửi cho admin để đăng ký\n"
                f"3. Admin sẽ thêm vào database\n"
                f"4. Chạy lại tool\n\n"
                f"Liên hệ admin để được hỗ trợ!"
            )
            root.destroy()
            return False
            
    except requests.exceptions.RequestException as e:
        print(f"⚠️ Không thể kết nối Google Sheets: {e}")
        messagebox.showwarning(
            "Lỗi Kết Nối",
            f"⚠️ Không thể kiểm tra license (lỗi mạng)\n\n"
            f"Lỗi: {e}\n\n"
            f"Vui lòng:\n"
            f"- Kiểm tra kết nối internet\n"
            f"- Thử lại sau"
        )
        root.destroy()
        return False
    except Exception as e:
        print(f"❌ Lỗi kiểm tra license: {e}")
        messagebox.showerror(
            "Lỗi",
            f"❌ Lỗi kiểm tra license:\n\n{e}\n\n"
            f"Liên hệ admin để được hỗ trợ."
        )
        root.destroy()
        return False

# ========= SETTINGS =========
def _resolve_settings_path():
    """Tìm settings.json: exe folder -> CWD -> script folder"""
    cand = []
    try:
        if getattr(sys, "frozen", False):
            cand.append(Path(sys.executable).parent / "settings.json")
    except:
        pass
    cand.append(Path.cwd() / "settings.json")
    try:
        cand.append(Path(__file__).parent / "settings.json")
    except:
        pass
    
    for p in cand:
        if p.is_file():
            return p
    
    try:
        if getattr(sys, "frozen", False):
            return Path(sys.executable).parent / "settings.json"
    except:
        pass
    return Path(__file__).parent / "settings.json"

SETTINGS_PATH = _resolve_settings_path()

REQUIRED_KEYS = [
    "excel_file", "download_dir", "img_done_dir", "result_xlsx",
    "chrome_exe", "user_data_dir", "profile_dir_name",
    "max_wait_dl", "quiet_seconds",
    "nsfw_h5",  # NSFW model path
    "gemini_api_keys",  # Danh sách API keys
    "gemini_prompt"  # Prompt cho Gemini
]

def load_settings(strict=True):
    """Load settings từ JSON"""
    if not SETTINGS_PATH.is_file():
        raise FileNotFoundError(f"Không tìm thấy settings.json tại: {SETTINGS_PATH}")
    
    with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
        s = json.load(f)
    
    if strict:
        missing = [k for k in REQUIRED_KEYS if k not in s]
        if missing:
            raise KeyError(f"Thiếu keys trong settings.json: {missing}")
    
    return s

# ========= EXCEL HELPERS =========

def read_image_links(excel_file):
    """Đọc các link ảnh từ file Excel (các cột Image 1..8)"""
    df = pd.read_excel(excel_file)
    image_cols = [col for col in df.columns if col.lower().startswith('image')]
    rows = []
    for idx, row in df.iterrows():
        links = [str(row[col]).strip() for col in image_cols if pd.notna(row[col]) and str(row[col]).strip()]
        product_url = str(row['Product URL']).strip() if 'Product URL' in df.columns and pd.notna(row['Product URL']) else ""
        product_name = str(row['Product Name']).strip() if 'Product Name' in df.columns and pd.notna(row['Product Name']) else ""
        if links:
            rows.append({'links': links, 'product_url': product_url, 'product_name': product_name})
    return rows

def excel_ensure(xlsx_path):
    """Tạo file Excel nếu chưa có"""
    p = Path(xlsx_path)
    if not p.is_file():
        p.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.append(["STT", "Product Name", "Link", "Ảnh số", "Link ảnh được chọn"])
        wb.save(xlsx_path)

def excel_next_stt(xlsx_path):
    """Lấy STT tiếp theo"""
    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active
        return ws.max_row
    except:
        return 1

def excel_append_row(xlsx_path, row_data):
    """Append row vào Excel"""
    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active
        ws.append(row_data)
        wb.save(xlsx_path)
    except Exception as e:
        print(f"⚠️ Lỗi ghi Excel: {e}")

# ========= DIRECTORY HELPERS =========
def ensure_dirs(*paths):
    """Tạo thư mục nếu chưa có"""
    for p in paths:
        Path(p).mkdir(parents=True, exist_ok=True)

def clear_download_dir(download_dir):
    """Xóa tất cả ảnh trong folder download"""
    deleted_count = 0
    failed_files = []
    
    for ext in ("*.jpg", "*.jpeg", "*.png", "*.webp", "*.gif", "*.bmp"):
        for f in glob.glob(os.path.join(download_dir, ext)):
            try:
                # Thử xóa nhiều lần nếu file bị lock
                for attempt in range(3):
                    try:
                        os.remove(f)
                        deleted_count += 1
                        break
                    except:
                        if attempt < 2:
                            time.sleep(0.2)  # Đợi file được release
                        else:
                            raise
            except Exception as e:
                failed_files.append(os.path.basename(f))
    
    if deleted_count > 0:
        print(f"🗑️ Đã xóa {deleted_count} ảnh")
    if failed_files:
        print(f"⚠️ Không xóa được {len(failed_files)} file (đang bị lock)")
    if deleted_count == 0 and not failed_files:
        print(f"✓ Folder trống")


def stable_key(path):
    """Sort key cho file (theo số nếu có)"""
    import re
    name = os.path.basename(path)
    nums = re.findall(r'\d+', name)
    return (int(nums[0]) if nums else 0, name)

# ========= NSFW DETECTOR =========
_NSFWO_G = None

def ensure_nsfw_model(model_path):
    """Load NSFW model 1 lần"""
    global _NSFWO_G
    if _NSFWO_G is None:
        print("🔍 Loading NSFW model...")
        _NSFWO_G = predict.load_model(model_path)
        print("✅ NSFW model loaded")
    return _NSFWO_G

def is_sensitive_image(img_path, model):
    """Kiểm tra ảnh NSFW"""
    try:
        result = predict.classify(model, img_path)
        for k, v in result.items():
            scores = v
            break
        
        porn = scores.get('porn', 0)
        hentai = scores.get('hentai', 0)
        sexy = scores.get('sexy', 0)
        
        return (porn > 0.5 or hentai > 0.5 or sexy > 0.6)
    except:
        return False

def _safe_read_image(path):
    """Đọc ảnh an toàn"""
    try:
        img = cv2.imread(path, cv2.IMREAD_COLOR)
        if img is None:
            return None
        return img
    except:
        return None

# ========= GEMINI AI =========
_GEMINI_MODEL = None
_CURRENT_API_KEY_INDEX = 0
_API_KEYS_LIST = []

def load_gemini_model(api_key):
    """Load Gemini model với API key cụ thể"""
    global _GEMINI_MODEL
    try:
        print(f"🔍 Loading Gemini Vision model (Key #{_CURRENT_API_KEY_INDEX + 1})...")
        genai.configure(api_key=api_key)
        _GEMINI_MODEL = genai.GenerativeModel('gemini-2.5-flash-lite')
        print("✅ Gemini 2.5 Flash lite loaded")
        return _GEMINI_MODEL
    except Exception as e:
        print(f"❌ Gemini failed: {e}")
        return None

def get_next_api_key():
    """Lấy API key tiếp theo trong danh sách"""
    global _CURRENT_API_KEY_INDEX, _API_KEYS_LIST
    
    if not _API_KEYS_LIST:
        return None
    
    _CURRENT_API_KEY_INDEX = (_CURRENT_API_KEY_INDEX + 1) % len(_API_KEYS_LIST)
    return _API_KEYS_LIST[_CURRENT_API_KEY_INDEX]

def init_api_keys(api_keys):
    """Khởi tạo danh sách API keys"""
    global _API_KEYS_LIST, _CURRENT_API_KEY_INDEX
    
    if isinstance(api_keys, str):
        _API_KEYS_LIST = [api_keys]
    elif isinstance(api_keys, list):
        _API_KEYS_LIST = api_keys
    else:
        _API_KEYS_LIST = []
    
    _CURRENT_API_KEY_INDEX = 0
    print(f"📝 Đã load {len(_API_KEYS_LIST)} API key(s)")

def is_quota_error(error_msg):
    """Kiểm tra xem có phải lỗi hết quota không"""
    quota_keywords = [
        "quota",
        "rate limit",
        "429",
        "RESOURCE_EXHAUSTED",
        "quota exceeded",
        "limit exceeded"
    ]
    error_lower = str(error_msg).lower()
    return any(kw.lower() in error_lower for kw in quota_keywords)

def pick_best_image_with_gemini(file_list, gemini_api_keys=None):
    """
    🤖 GEMINI AI - Chọn ảnh tốt nhất cho Veo3
    
    Gửi TẤT CẢ ảnh cho Gemini, AI sẽ:
    1. Phân tích sản phẩm đang bán
    2. Loại bỏ ảnh xấu (infographic, hướng dẫn, đồ lót...)
    3. Chọn 1 ảnh sản phẩm TỐT NHẤT
    
    Tự động đổi API key khi hết quota.
    
    Return: (best_path, best_index) hoặc (None, -1)
    """
    global _NSFWO_G, _API_KEYS_LIST, _CURRENT_API_KEY_INDEX
    
    print(f"\n🤖 GEMINI AI - Đang phân tích {len(file_list)} ảnh...")
    
    # BƯỚC 1: NSFW Check + Quality Filter
    safe_files = []
    for p in file_list:
        if is_sensitive_image(p, _NSFWO_G):
            print(f"🚫 Ảnh nhạy cảm (bỏ qua): {os.path.basename(p)}")
            continue
        
        try:
            img = _safe_read_image(p)
            if img is None:
                continue
            h, w = img.shape[:2]
            
            # Độ phân giải tối thiểu
            if min(h, w) < 600:
                print(f"🚫 Độ phân giải thấp ({w}x{h}): {os.path.basename(p)}")
                continue
            
            # Aspect ratio hợp lý
            aspect = w / h
            if aspect < 0.5 or aspect > 2.5:
                print(f"🚫 Tỷ lệ khung hình lệch ({aspect:.2f}): {os.path.basename(p)}")
                continue
            
            safe_files.append(p)
        except:
            continue
    
    if not safe_files:
        print("❌ Không có ảnh hợp lệ sau khi lọc")
        return None, -1
    
    print(f"✅ Còn {len(safe_files)} ảnh hợp lệ, gửi cho Gemini...")
    
    # BƯỚC 2: GỬI CHO GEMINI (với retry khi hết quota)
    max_retries = len(_API_KEYS_LIST) if _API_KEYS_LIST else 1
    
    # Prompt lấy từ settings
    # Sử dụng SETTINGS_PATH là biến toàn cục
    global SETTINGS_PATH
    try:
        with open(SETTINGS_PATH, "r", encoding="utf-8") as f:
            s = json.load(f)
        prompt = s.get("gemini_prompt", "")
    except:
        prompt = ""
    if not prompt:
        prompt = f"Bạn là chuyên gia chọn ảnh cho video quảng cáo Veo3. Tôi có {len(safe_files)} ảnh sản phẩm từ Shopee.\n..."
    
    # Load tất cả ảnh (copy vào memory để tránh file lock)
    images = []
    for p in safe_files:
        try:
            with PILImage.open(p) as img:
                # Copy ảnh vào memory để đóng file ngay
                img_copy = img.copy()
                images.append(img_copy)
        except:
            continue
    
    if not images:
        print("❌ Không thể load ảnh")
        return None, -1
    
    # Retry loop với nhiều API keys
    for retry in range(max_retries):
        try:
            # Lấy API key hiện tại
            current_key = _API_KEYS_LIST[_CURRENT_API_KEY_INDEX] if _API_KEYS_LIST else gemini_api_keys
            
            model = load_gemini_model(current_key)
            if model is None:
                print("❌ Gemini không khả dụng")
                if retry < max_retries - 1:
                    get_next_api_key()
                    continue
                return None, -1
            
            print(f"📤 Đang gửi {len(images)} ảnh cho Gemini AI...")
            
            # Gửi cho Gemini
            response = model.generate_content([prompt] + images)
            result_text = response.text.strip()
            
            print(f"\n🤖 Gemini phản hồi:\n{result_text}\n")
            
            # Parse JSON
            import re
            json_match = re.search(r'\{[^}]+\}', result_text, re.DOTALL)
            if not json_match:
                # Fallback: tìm số
                index_match = re.search(r'"best_image_index"\s*:\s*(\d+)', result_text)
                if index_match:
                    best_idx = int(index_match.group(1))
                else:
                    print("❌ Không parse được JSON")
                    return None, -1
            else:
                result_json = json.loads(json_match.group(0))
                best_idx = result_json.get('best_image_index', -1)
                product_type = result_json.get('product_type', 'Unknown')
                reason = result_json.get('reason', '')

                print(f"📦 Sản phẩm: {product_type}")
                print(f"💡 Lý do: {reason}")

                # Kiểm tra REJECT
                if "REJECT_SENSITIVE" in reason or "No good" in reason or best_idx == -1:
                    print(f"🚫 Gemini từ chối: {reason}")
                    return None, -1

            # Validate index
            if best_idx < 0 or best_idx >= len(safe_files):
                print(f"❌ Index không hợp lệ: {best_idx}")
                return None, -1

            best_path = safe_files[best_idx]

            # Tìm index trong file_list gốc
            original_idx = file_list.index(best_path) + 1  # +1 vì Excel đếm từ 1

            print(f"✅ Gemini chọn: {os.path.basename(best_path)} (#{original_idx})")

            return best_path, original_idx
            
        except Exception as e:
            error_msg = str(e)
            print(f"❌ Gemini error: {error_msg}")
            
            # Kiểm tra xem có phải lỗi hết quota không
            if is_quota_error(error_msg):
                print(f"⚠️ API Key #{_CURRENT_API_KEY_INDEX + 1} hết quota!")
                
                # Nếu còn key khác thì thử tiếp
                if retry < max_retries - 1:
                    next_key = get_next_api_key()
                    print(f"🔄 Đổi sang API Key #{_CURRENT_API_KEY_INDEX + 1}...")
                    time.sleep(1)
                    continue
                else:
                    print(f"❌ Tất cả {len(_API_KEYS_LIST)} API keys đều hết quota!")
                    return None, -1
            else:
                # Lỗi khác thì không retry
                import traceback
                traceback.print_exc()
                return None, -1
    
    # Hết retry
    print("❌ Không thể kết nối Gemini sau nhiều lần thử")
    return None, -1

# ========= RUNNER THREAD =========
class RunnerThread(threading.Thread):
    def __init__(self, ui):
        super().__init__(daemon=True)
        self.ui = ui
        self._stop_event = threading.Event()
        self.captcha_event = threading.Event()
    
    def stop(self):
        self._stop_event.set()
    

    def download_images(self, links, download_dir):
        """Tải các link ảnh về download_dir, trả về danh sách file đã tải"""
        files = []
        for idx, url in enumerate(links, start=1):
            try:
                ext = os.path.splitext(url)[-1]
                if ext.lower() not in ['.jpg', '.jpeg', '.png', '.webp', '.bmp', '.gif']:
                    ext = '.jpg'
                out_path = os.path.join(download_dir, f"img_{idx}{ext}")
                resp = requests.get(url, timeout=15)
                if resp.status_code == 200:
                    with open(out_path, 'wb') as f:
                        f.write(resp.content)
                    files.append(out_path)
                else:
                    self.ui.log(f"  ⚠️ Không tải được ảnh: {url}")
            except Exception as e:
                self.ui.log(f"  ⚠️ Lỗi tải ảnh: {url} ({e})")
        return files

    def run(self):
        s = self.ui.settings
        ensure_dirs(s["download_dir"], s["img_done_dir"])
        excel_ensure(s["result_xlsx"])

        try:
            all_image_rows = read_image_links(s["excel_file"])
        except Exception as e:
            self.ui.log(f"❌ Không đọc được file {s['excel_file']}: {e}")
            self.ui.after(0, self.ui.on_thread_done)
            return

        self.ui.log(f"Tổng {len(all_image_rows)} sản phẩm.")


        for idx_row, row in enumerate(all_image_rows, start=1):
            if self._stop_event.is_set():
                break

            links = row['links']
            product_url = row['product_url']
            product_name = row['product_name']

            clear_download_dir(s["download_dir"])
            self.ui.log(f"\n[{idx_row}] Đang tải {len(links)} ảnh...")

            files = self.download_images(links, s["download_dir"])
            # Hiển thị ảnh vừa tải về trên giao diện
            stt = excel_next_stt(s["result_xlsx"])

            if not files:
                self.ui.log("  ❌ Không tải được ảnh nào.")
                excel_append_row(s["result_xlsx"], [stt, product_name, product_url, 0, ""])
                continue

            files = sorted(files, key=stable_key)

            # Chỉ còn chế độ Gemini AI
            self.ui.log("🤖 Gemini đang chọn ảnh tốt nhất...")
            best_path, best_idx = pick_best_image_with_gemini(files)
            if not best_path:
                self.ui.log("  ❌ Gemini không chọn được ảnh.")
                excel_append_row(s["result_xlsx"], [stt, product_name, product_url, 0, ""])
            else:
                try:
                    im = Image.open(best_path).convert("RGB")
                    self.ui.show_preview_image(im)
                except Exception as e:
                    self.ui.log(f"  ⚠️ Lỗi preview: {e}")
                self.ui.log(f"  ✅ Chọn ảnh #{best_idx}")
                # Lấy link gốc từ file Excel
                selected_link = links[best_idx-1] if best_idx > 0 and best_idx <= len(links) else ""
                excel_append_row(s["result_xlsx"], [stt, product_name, product_url, best_idx, selected_link])

            clear_download_dir(s["download_dir"])
            time.sleep(0.5)

        self.ui.log("\n✅ Hoàn tất!")
        self.ui.after(0, self.ui.on_thread_done)

# ========= GUI =========
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        
        # DPI scaling
        try:
            if sys.platform.startswith("win"):
                import ctypes
                ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except:
            pass
        
        self.title("LỌC ẢNH AI QUANG MINH")
        self.geometry("900x600")
        self.minsize(900, 800)
        self.configure(bg="#eaf0fa")
        
        # Load settings
        self.settings = load_settings(strict=True)
        
        # Khởi tạo API keys
        api_keys = self.settings.get("gemini_api_keys") or self.settings.get("gemini_api_key")
        init_api_keys(api_keys)
        
        # Pre-load NSFW model
        ensure_nsfw_model(self.settings["nsfw_h5"])
        
        # Mode selection variable - PHẢI khởi tạo TRƯỚC build_ui()
        self.pick_mode = tk.IntVar(value=1)  # Default: Gemini AI
        
        self.runner = None
        
        # Build UI
        self.build_ui()
    
    def build_ui(self):
        # ============ HEADER ============
        header = tk.Frame(self, bg="#3a5a99", height=70)
        header.pack(side=tk.TOP, fill=tk.X)

        # Gradient effect (simulate)
        gradient = tk.Canvas(header, height=70, bg="#3a5a99", highlightthickness=0)
        gradient.pack(fill=tk.BOTH, expand=True)
        for i in range(70):
            color = f"#%02x%02x%02x" % (58 + i//2, 90 + i, 153 + i)
            gradient.create_rectangle(0, i, 900, i+1, outline=color, fill=color)

        # Title with icon and shadow
        title_frame = tk.Frame(header, bg="#3a5a99", height=70)
        title_frame.place(relx=0.5, rely=0.5, anchor=tk.CENTER)

        tk.Label(
            title_frame,
            text="✨ LỌC ẢNH AI QUANG MINH ✨",
            bg="#3a5a99",
            fg="#fff",
            font=("Montserrat", 20, "bold"),
            pady=4
        ).pack()

        tk.Label(
            title_frame,
            text="AI Quang Minh | Lọc ảnh Shopee | Tự động chọn ảnh đẹp | Excel Automation",
            bg="#3a5a99",
            fg="#eaf0fa",
            font=("Montserrat", 10, "italic")
        ).pack()

        # ============ MODE SELECTION ============
        mode_container = tk.Frame(self, bg="#eaf0fa")
        mode_container.pack(side=tk.TOP, fill=tk.X, padx=12, pady=8)

        mode_frame = tk.Frame(mode_container, bg="#fff", relief=tk.RIDGE, borderwidth=0, highlightbackground="#b8c1ec", highlightthickness=2)
        mode_frame.pack(fill=tk.X, padx=4, pady=3)

        # Mode header
        mode_header = tk.Frame(mode_frame, bg="#4f6cb3", height=32)
        mode_header.pack(fill=tk.X)

        tk.Label(
            mode_header,
            text="⚙️ CHẾ ĐỘ CHỌN ẢNH",
            bg="#4f6cb3",
            fg="#fff",
            font=("Montserrat", 11, "bold")
        ).pack(side=tk.LEFT, padx=12, pady=6)

        # Mode options
        mode_options = tk.Frame(mode_frame, bg="#fff")
        mode_options.pack(fill=tk.X, padx=12, pady=8)

        # Option Gemini AI only
        opt1_frame = tk.Frame(mode_options, bg="#eaf7ea", relief=tk.GROOVE, borderwidth=0, highlightbackground="#4f6cb3", highlightthickness=2)
        opt1_frame.pack(side=tk.LEFT, padx=8, ipadx=12, ipady=8)

        self.radio_mode1 = tk.Radiobutton(
            opt1_frame,
            text="🤖 AI Quang Minh",
            variable=self.pick_mode,
            value=1,
            font=("Montserrat", 10, "bold"),
            bg="#eaf7ea",
            fg="#2e7d32",
            selectcolor="#c8e6c9",
            activebackground="#eaf7ea",
            cursor="hand2"
        )
        self.radio_mode1.pack(anchor=tk.W)
        self.radio_mode1.bind("<Enter>", lambda e: self.radio_mode1.config(bg="#d0f0c0"))
        self.radio_mode1.bind("<Leave>", lambda e: self.radio_mode1.config(bg="#eaf7ea"))

        tk.Label(
            opt1_frame,
            text="✓ AI chọn ảnh đẹp nhất\n✓ Lọc ảnh xấu tự động\n✓ Tối ưu cho Shopee",
            bg="#eaf7ea",
            fg="#555",
            font=("Montserrat", 9),
            justify=tk.LEFT
        ).pack(anchor=tk.W, padx=12)

        # ============ CONTROL PANEL ============
        ctrl_container = tk.Frame(self, bg="#eaf0fa")
        ctrl_container.pack(side=tk.TOP, fill=tk.X, padx=12, pady=(0, 8))

        ctrl = tk.Frame(ctrl_container, bg="#fff", relief=tk.RIDGE, borderwidth=0, highlightbackground="#b8c1ec", highlightthickness=2)
        ctrl.pack(fill=tk.X, padx=4, pady=3)

        ctrl_inner = tk.Frame(ctrl, bg="#fff")
        ctrl_inner.pack(padx=12, pady=12)

        # --- Cài đặt Gemini & Prompt ---
        settings_frame = tk.LabelFrame(ctrl_inner, text="Cài đặt AI Quang Minh", bg="#fff", font=("Montserrat", 10, "bold"), fg="#3a5a99", relief=tk.RIDGE, borderwidth=0, highlightbackground="#4f6cb3", highlightthickness=2)
        settings_frame.pack(side=tk.LEFT, padx=8, pady=4, fill=tk.Y)

        tk.Label(settings_frame, text="API Keys (mỗi key một dòng):", bg="#fff", anchor="w", font=("Montserrat", 9)).pack(fill=tk.X, padx=4, pady=(4,0))
        self.api_keys_text = tk.Text(settings_frame, width=50, height=4, font=("Consolas", 9), relief=tk.GROOVE, borderwidth=2)
        self.api_keys_text.pack(padx=4, pady=2)
        keys = self.settings.get("gemini_api_keys", [])
        self.api_keys_text.insert("1.0", "\n".join(keys))
        self.api_keys_text.bind("<Enter>", lambda e: self.api_keys_text.config(bg="#f0f8ff"))
        self.api_keys_text.bind("<Leave>", lambda e: self.api_keys_text.config(bg="#fff"))

        tk.Label(settings_frame, text="Prompt cho Gemini AI:", bg="#fff", anchor="w", font=("Montserrat", 9)).pack(fill=tk.X, padx=4, pady=(8,0))
        self.prompt_text = tk.Text(settings_frame, width=50, height=8, font=("Consolas", 9), relief=tk.GROOVE, borderwidth=2)
        self.prompt_text.pack(padx=4, pady=2)
        prompt_val = self.settings.get("gemini_prompt", "")
        self.prompt_text.insert("1.0", prompt_val)
        self.prompt_text.bind("<Enter>", lambda e: self.prompt_text.config(bg="#f0f8ff"))
        self.prompt_text.bind("<Leave>", lambda e: self.prompt_text.config(bg="#fff"))

        self.save_settings_btn = tk.Button(settings_frame, text="💾 Lưu cài đặt", bg="#3a5a99", fg="#fff", font=("Montserrat", 10, "bold"), command=self.save_settings, relief=tk.RAISED, borderwidth=0, cursor="hand2")
        self.save_settings_btn.pack(padx=4, pady=6, fill=tk.X)
        self.save_settings_btn.bind("<Enter>", lambda e: self.save_settings_btn.config(bg="#4f6cb3"))
        self.save_settings_btn.bind("<Leave>", lambda e: self.save_settings_btn.config(bg="#3a5a99"))

        # Buttons
        btn_frame = tk.Frame(ctrl_inner, bg="#fff")
        btn_frame.pack(side=tk.LEFT, padx=6)

        self.btn_start = tk.Button(
            btn_frame,
            text="▶ BẮT ĐẦU",
            bg="#43b581",
            fg="#fff",
            font=("Montserrat", 11, "bold"),
            width=10,
            height=1,
            relief=tk.RAISED,
            borderwidth=0,
            cursor="hand2",
            command=self.on_start
        )
        self.btn_start.pack(side=tk.LEFT, padx=6)
        self.btn_start.bind("<Enter>", lambda e: self.btn_start.config(bg="#2ecc71"))
        self.btn_start.bind("<Leave>", lambda e: self.btn_start.config(bg="#43b581"))

        self.btn_stop = tk.Button(
            btn_frame,
            text="⏹ DỪNG LẠI",
            bg="#e74c3c",
            fg="#fff",
            font=("Montserrat", 11, "bold"),
            width=10,
            height=1,
            relief=tk.RAISED,
            borderwidth=0,
            cursor="hand2",
            command=self.on_stop,
            state=tk.DISABLED
        )
        self.btn_stop.pack(side=tk.LEFT, padx=6)
        self.btn_stop.bind("<Enter>", lambda e: self.btn_stop.config(bg="#c0392b") if self.btn_stop['state'] == 'normal' else None)
        self.btn_stop.bind("<Leave>", lambda e: self.btn_stop.config(bg="#e74c3c") if self.btn_stop['state'] == 'normal' else None)

        # Info panel
        info_frame = tk.Frame(ctrl_inner, bg="#f7e6e6", relief=tk.RIDGE, borderwidth=0, highlightbackground="#b8c1ec", highlightthickness=2)
        info_frame.pack(side=tk.LEFT, padx=12, fill=tk.BOTH, expand=True)

        key_count = len(_API_KEYS_LIST) if _API_KEYS_LIST else 1

        tk.Label(
            info_frame,
            text=f"📁 File Excel",
            bg="#f7e6e6",
            fg="#3a5a99",
            font=("Montserrat", 10, "bold")
        ).pack(anchor=tk.W, padx=8, pady=(8, 0))

        tk.Label(
            info_frame,
            text=os.path.basename(self.settings['excel_file']),
            bg="#f7e6e6",
            fg="#c0392b",
            font=("Montserrat", 12, "bold")
        ).pack(anchor=tk.W, padx=8)

        tk.Label(
            info_frame,
            text=f"🔑 API Keys: {key_count}",
            bg="#f7e6e6",
            fg="#3a5a99",
            font=("Montserrat", 10)
        ).pack(anchor=tk.W, padx=8, pady=(6, 8))

        # ============ MAIN CONTENT ============
        content = tk.Frame(self, bg="#eaf0fa")
        content.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=12, pady=(0, 10))

        # Left: Log panel
        left_container = tk.Frame(content, bg="#fff", relief=tk.RIDGE, borderwidth=0, highlightbackground="#b8c1ec", highlightthickness=2)
        left_container.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(6, 10))

        log_header = tk.Frame(left_container, bg="#4f6cb3", height=32)
        log_header.pack(fill=tk.X)

        tk.Label(
            log_header,
            text="📋 NHẬT KÝ XỬ LÝ",
            bg="#4f6cb3",
            fg="#fff",
            font=("Montserrat", 10, "bold")
        ).pack(side=tk.LEFT, padx=12, pady=6)

        log_content = tk.Frame(left_container, bg="#fff")
        log_content.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        self.log_text = tk.Text(
            log_content,
            wrap=tk.WORD,
            font=("Consolas", 9),
            bg="#f7fafd",
            fg="#3a5a99",
            relief=tk.GROOVE,
            borderwidth=2,
            padx=6,
            pady=6
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scroll = tk.Scrollbar(log_content, command=self.log_text.yview)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=scroll.set)

        # Right: Preview panel
        right_container = tk.Frame(content, bg="#fff", relief=tk.RIDGE, borderwidth=0, highlightbackground="#b8c1ec", highlightthickness=2, width=220)
        right_container.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 6))
        right_container.pack_propagate(False)

        preview_header = tk.Frame(right_container, bg="#4f6cb3", height=32)
        preview_header.pack(fill=tk.X)

        tk.Label(
            preview_header,
            text="🖼️ XEM TRƯỚC ẢNH",
            bg="#4f6cb3",
            fg="#fff",
            font=("Montserrat", 10, "bold")
        ).pack(side=tk.LEFT, padx=12, pady=6)

        preview_content = tk.Frame(right_container, bg="#f7fafd")
        preview_content.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        self.preview_label = tk.Label(
            preview_content,
            bg="#f7e6e6",
            text="Chưa có ảnh",
            fg="#3a5a99",
            font=("Montserrat", 10),
            relief=tk.GROOVE,
            borderwidth=2
        )
        self.preview_label.pack(fill=tk.BOTH, expand=True)

    def log(self, msg):
        """Append log"""
        def _do():
            self.log_text.insert(tk.END, msg + "\n")
            self.log_text.see(tk.END)
        
        if threading.current_thread() != threading.main_thread():
            self.after(0, _do)
        else:
            _do()
    
    def show_preview_image(self, pil_img):
        """Show preview"""
        def _do():
            w, h = pil_img.size
            max_w, max_h = 280, 500
            scale = min(max_w / w, max_h / h, 1.0)
            new_w = int(w * scale)
            new_h = int(h * scale)
            resized = pil_img.resize((new_w, new_h), Image.Resampling.LANCZOS)
            photo = ImageTk.PhotoImage(resized)
            self.preview_label.config(image=photo, text="")
            self.preview_label.image = photo
        
        if threading.current_thread() != threading.main_thread():
            self.after(0, _do)
        else:
            _do()
    
    def on_start(self):
        """Start runner"""
        if self.runner and self.runner.is_alive():
            messagebox.showwarning("Cảnh báo", "Đang chạy rồi!")
            return
        
        # Khởi tạo và bắt đầu runner
        self.runner = RunnerThread(self)
        self.runner.start()
        
        # Disable UI elements
        self.btn_start.config(state=tk.DISABLED)
        self.btn_stop.config(state=tk.NORMAL)
        self.radio_mode1.config(state=tk.DISABLED)
        
    def save_settings(self):
        """Lưu lại API keys và prompt vào settings.json"""
        import json
        from tkinter import messagebox
        try:
            keys = [k.strip() for k in self.api_keys_text.get("1.0", tk.END).splitlines() if k.strip()]
            prompt_val = self.prompt_text.get("1.0", tk.END).strip()
            self.settings["gemini_api_keys"] = keys
            self.settings["gemini_prompt"] = prompt_val
            # Lưu file
            with open(str(SETTINGS_PATH), "w", encoding="utf-8") as f:
                json.dump(self.settings, f, indent=2, ensure_ascii=False)
            # Reload các biến liên quan
            init_api_keys(keys)
            self.log(f"Đã lưu cài đặt!")
            messagebox.showinfo("Thành công", "Đã lưu cài đặt vào settings.json!")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể lưu: {e}")
        self.log_text.delete("1.0", tk.END)
    
    def on_stop(self):
        """Stop runner"""
        if self.runner:
            self.runner.stop()
        self.log("\n⏹ Đã dừng.")
        self.on_thread_done()
    
    def on_thread_done(self):
        """Thread finished"""
        self.btn_start.config(state=tk.NORMAL)
        self.btn_stop.config(state=tk.DISABLED)
        self.radio_mode1.config(state=tk.NORMAL)
    
    def show_captcha_dialog(self):
        """Hiển thị dialog khi phát hiện captcha"""
        messagebox.showinfo(
            "Captcha Detected",
            "🚫 Phát hiện captcha!\n\n"
            "Vui lòng giải captcha thủ công trong trình duyệt.\n"
            "Sau khi giải xong, nhấn OK để tool tiếp tục."
        )
        if self.runner:
            self.runner.captcha_event.set()

# ========= ADMIN TOOLS =========
def show_machine_id():
    """Tool hiển thị Machine ID - Dành cho khách hàng"""
    machine_id = _get_machine_id()
    
    print("\n" + "="*70)
    print("🖥️ MACHINE ID CỦA MÁY NÀY")
    print("="*70)
    print(f"\nMachine ID: {machine_id}")
    print("\n📋 Hướng dẫn:")
    print("1. Copy Machine ID này")
    print("2. Gửi cho admin để đăng ký license")
    print("3. Đợi admin thêm vào database")
    print("4. Chạy lại tool")
    print("\n⚠️ Mỗi máy có Machine ID khác nhau!")
    print("="*70 + "\n")
    
    # Hiển thị popup
    root = tk.Tk()
    root.withdraw()  # Ẩn cửa sổ chính
    messagebox.showinfo(
        "Machine ID",
        f"🖥️ MACHINE ID CỦA MÁY NÀY\n\n"
        f"{machine_id}\n\n"
        f"📋 Hướng dẫn:\n"
        f"1. Copy Machine ID này\n"
        f"2. Gửi cho admin để đăng ký\n"
        f"3. Chạy lại tool sau khi được kích hoạt"
    )
    root.destroy()

# ========= MAIN =========
if __name__ == "__main__":
    # Tool mode: Hiển thị Machine ID
    if len(sys.argv) > 1 and sys.argv[1] == "--show-id":
        show_machine_id()
        sys.exit(0)
    
    # Kiểm tra license từ Google Sheets TRƯỚC KHI khởi động
    if not check_license_from_google_sheet():
        print("\n❌ Tool bị chặn do chưa có license!")
        print(f"Machine ID: {_get_machine_id()}")
        print("Gửi Machine ID này cho admin để được kích hoạt.")
        sys.exit(1)
    
    try:
        app = App()
        app.mainloop()
    except Exception as e:
        messagebox.showerror("Lỗi", f"Lỗi khởi động:\n{e}")
        import traceback
        traceback.print_exc()
